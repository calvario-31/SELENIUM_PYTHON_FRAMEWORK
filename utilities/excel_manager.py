from pathlib import Path

from openpyxl import load_workbook


def get_excel_sheet(excel_relative_path, sheet_name):
    excel_complete_path = Path("resources/data/" + excel_relative_path)
    excel_workbook = load_workbook(filename=excel_complete_path)
    return excel_workbook[sheet_name]


def get_dictionary_from_key(excel_sheet, key_name):
    dictionary = {}
    for i in range(1, excel_sheet.max_row + 1):
        if excel_sheet.cell(row=i, column=1).value == key_name:
            for j in range(2, excel_sheet.max_column + 1):
                dictionary[excel_sheet.cell(row=1, column=j).value] = excel_sheet.cell(row=i, column=j).value
    return dictionary


def get_all_rows(excel_sheet):
    rows_list = []
    for i in range(2, excel_sheet.max_row + 1):
        dictionary_row = {}
        for j in range(1, excel_sheet.max_column):
            dictionary_row[excel_sheet.cell(row=1, column=j).value] = excel_sheet.cell(row=i, column=j).value

        rows_list.append(dictionary_row)

    return rows_list
